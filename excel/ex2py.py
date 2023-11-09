from openpyxl import load_workbook
from  excel import PyTemp
import re
# import sys
import os


RegBankMes = {}
RegMap = {}

def ReadExcel(input_path, output_path):
    # if len(sys.argv)>=2:    input_path = sys.argv[1]
    # else:   raise Exception("Input file not exist!")
    # if len(sys.argv)>=3:    output_path = sys.argv[2]
    # else:   output_path = ''

    workbook = load_workbook(input_path)
    regBank = workbook["REGBANK"]
    row = regBank.max_row
    column = regBank.max_column
    tableLine = -1
    regName = ''
    filedName = ''
    for i in range(row):
        iNum = 1 + i
        if regBank['A' + str(iNum)].value == 'Name':
            RegBankMes['name'] = regBank['B' + str(iNum)].value
        elif regBank['A' + str(iNum)].value == 'Size (in KB)':
            RegBankMes['size'] = regBank['B' + str(iNum)].value
        elif regBank['A' + str(iNum)].value == 'Software Interface':
            RegBankMes['interface'] = str.lower(regBank['B' + str(iNum)].value)
        elif regBank['A' + str(iNum)].value == 'Interface Width':
            RegBankMes['width'] = regBank['B' + str(iNum)].value
        elif regBank['A' + str(iNum)].value == 'Description':
            RegBankMes['description'] = regBank['B' + str(iNum)].value
        elif regBank['A' + str(iNum)].value == 'Check':
            RegBankMes['check'] = str(regBank['B' + str(iNum)].value).lower()
        elif regBank['A' + str(iNum)].value == 'RegName':
            tableLine = iNum
        elif iNum > tableLine and tableLine != -1:
            for j in range(column):
                jChar = chr(ord('A') + j)
                cellName = jChar + str(tableLine)
                cellValue = regBank[jChar + str(iNum)].value 
                if cellValue is not None and regBank[cellName].value  == 'RegName':
                    regName = cellValue
                    RegMap[regName] = {}
                elif cellValue is not None and regBank[cellName].value  == 'OffsetAddress':
                    # RegMap[regName]['OffsetAddress'] = int(cellValue,16)   # byte
                    RegMap[regName]['OffsetAddress'] = cellValue   # byte
                elif cellValue is not None and regBank[cellName].value  == 'RegType':
                    RegMap[regName]['RegType'] = cellValue
                elif cellValue is not None and regBank[cellName].value  == 'FieldName':
                    if 'Field' not in RegMap[regName]:
                        RegMap[regName]['Field'] = {}
                    filedName = cellValue
                    RegMap[regName]['Field'][filedName] = {}
                elif cellValue is not None and regBank[cellName].value  == 'Position':
                    End = cellValue[cellValue.index('[')+1:cellValue.index(':')]
                    Start = cellValue[cellValue.index(':')+1:cellValue.index(']')]
                    if 'Field' not in RegMap[regName]:
                        RegMap[regName]['Position'] = cellValue
                        RegMap[regName]['offset'] = Start
                        RegMap[regName]['bit'] = int(End) - int(Start) + 1
                    else:
                        RegMap[regName]['Field'][filedName]['Position'] = cellValue
                        RegMap[regName]['Field'][filedName]['offset'] = Start
                        RegMap[regName]['Field'][filedName]['bit'] = int(End) - int(Start) + 1
                elif cellValue is not None and regBank[cellName].value  == 'FieldType':
                    RegMap[regName]['Field'][filedName]['FieldType'] = cellValue
                elif cellValue is not None and regBank[cellName].value  == 'SoftwareAccess':
                    RegMap[regName]['Field'][filedName]['SoftwareAccess'] = cellValue
                elif cellValue is not None and regBank[cellName].value  == 'HardwareAccess':
                    RegMap[regName]['Field'][filedName]['HardwareAccess'] = cellValue
                elif cellValue is not None and regBank[cellName].value  == 'DefaultValue':
                    matcher = re.match(r'.*\'(\w)(\d*)',str(cellValue)) 
                    if matcher == None:
                        initValue = cellValue
                    elif matcher.group(1) == 'b':
                        initValue = '0b'+ matcher.group(2)
                    elif matcher.group(1) == 'h':
                        initValue = '0x'+matcher.group(2)
                    elif matcher.group(1) == 'd':
                        initValue = matcher.group(2)
                    if 'Field' not in RegMap[regName]:
                        RegMap[regName]['initValue'] = initValue
                    else:
                        RegMap[regName]['Field'][filedName]['initValue'] = initValue
                elif cellValue is not None and regBank[cellName].value  == 'Description':
                    if 'Field' not in RegMap[regName]:
                        RegMap[regName]['Description'] = cellValue
                    else:
                        RegMap[regName]['Field'][filedName]['Description'] = cellValue
                elif cellValue is not None and regBank[cellName].value  == 'LockDep':
                    if 'Field' not in RegMap[regName]:
                         RegMap[regName]['LockDep'] = cellValue.split(',')
                    else:
                        RegMap[regName]['Field'][filedName]['LockDep'] = cellValue.split(',')
                elif cellValue is not None and regBank[cellName].value  == 'MagicNumberDep':
                    RegMap[regName]['MagicNumberDep'] = cellValue.split(',')
                elif cellValue is not None and regBank[cellName].value  == 'MagicValue':
                    RegMap[regName]['MagicValue'] = cellValue

    PrintLog()
    
def PrintLog():
    datalog = open("datalog.txt",'w',encoding="utf-8")  
    for k,v in RegMap.items():
        print(k,file=datalog)
        print(v,file=datalog)
    for k,v in RegBankMes.items():
        print(k,file=datalog)
        print(v,file=datalog)


def CreatPy(input_path, output_path):
    
    ReadExcel(input_path, output_path)

    # if len(sys.argv)>=3:    output_path = sys.argv[2]
    # else:   output_path = ''

    pyCode = PyTemp.Head.replace('{name}',RegBankMes['name']).replace('{size}',str(RegBankMes['size'])).replace('{description}',RegBankMes['description']).replace('{width}',str(RegBankMes['width'])).replace('{interface}',RegBankMes['interface'])
    
    for index,regName in enumerate(RegMap):
        
        pyCode += '\n################################'+regName+'#######################################\n'
        
    pyCode = PyTemp.Head.replace('{name}',RegBankMes['name']).replace('{size}',str(RegBankMes['size'])).replace('{description}',RegBankMes['description']).replace('{width}',str(RegBankMes['width'])).replace('{interface}',RegBankMes['interface'])
    
    for index,regName in enumerate(RegMap):
        
        pyCode += '\n################################'+regName+'#######################################\n'
        pyCode += PyTemp.Reg.replace('{cnt}',str(index)).replace('{name}',regName).replace('{Description}',RegMap[regName].get('Description','')).replace('{RegType}',RegMap[regName]['RegType'])
        
        if RegMap[regName]['RegType'] == 'Normal':
            for fieldName,field in RegMap[regName]['Field'].items():
                RegCfg = PyTemp.RegCfg
                if field['FieldType'] == 'External':
                    RegCfg = RegCfg.replace('{Field}','External')
                else:
                    RegCfg = RegCfg.replace('{Field}','')
                if 'LockDep' not in  field:
                    RegCfg = RegCfg.replace(',lock_list={lockList}','')
                else:
                    lockList = ''
                    for item in field['LockDep']:
                        lockList += '\"' + str(item[:item.find('.')]) + '.' + str(item[item.find('.') + 1 :]) + '\",'
                    RegCfg = RegCfg.replace('{lockList}','[' + lockList[:-1] + ']')
                pyCode += RegCfg.replace('{cnt}',str(index)).replace('{name}',fieldName).replace('{bit}',str(field['bit'])).replace('{SoftwareAccess}',field['SoftwareAccess']).replace('{HardwareAccess}',field.get('HardwareAccess','Null')).replace('{initValue}',str(field['initValue'])).replace('{description}',field.get('Description','')).replace('{offset}',str(field['offset']))
                
        elif RegMap[regName]['RegType'] == 'Magic':
            pyCode += PyTemp.MagicRegCfg.replace('{cnt}',str(index)).replace('{MagicValue}',RegMap[regName]['MagicValue']).replace('{initValue}',RegMap[regName]['initValue']).replace('{bit}',str(RegMap[regName]['bit']))
        elif RegMap[regName]['RegType'] == 'Lock':
            for fieldName,field in RegMap[regName]['Field'].items():
                pyCode += PyTemp.LockRefCfg.replace('{cnt}',str(index)).replace('{bit}',str(field['bit'])).replace('{name}',fieldName).replace('{bit}',str(field['bit'])).replace('{description}',field.get('Description','')).replace('{offset}',str(field['offset']))
                
        ADD = PyTemp.ADD
        if 'MagicNumberDep' not in RegMap[regName]:
            ADD = ADD.replace(',magic_list={magicList}','')
        else:
            magicList = ''
            for item in RegMap[regName]['MagicNumberDep']:
                magicList += '\"' + str(item) + '\",'
            ADD = ADD.replace('{magicList}','[' + magicList[:-1] + ']')
        if 'LockDep' not in  RegMap[regName]:
            ADD = ADD.replace(',lock_list={lockList}','')
        else:
            LockList = ''
            for item in RegMap[regName]['LockDep']:
                LockList += '\"' + str(item[:item.find('.')]) + '.' + str(item[item.find('.') + 1 :]) + '\",'
            ADD = ADD.replace('{lockList}','[' + LockList[:-1] + ']')
        pyCode += ADD.replace('{cnt}',str(index)).replace('{OffsetAddress}',str(RegMap[regName]['OffsetAddress'])).replace('{name}',str(regName))

    pyCode += PyTemp.Gen.replace('{name}',RegBankMes['name'])
    
    if RegBankMes['check'] == 'true':
        pyCode += PyTemp.Check.replace('{name}',RegBankMes['name'])
        
    filename = RegBankMes['name']+"_rf_gen.py"
    if output_path != '': os.makedirs(output_path, exist_ok=True)
    output_file = os.path.join(output_path, filename)
    with open(output_file,'w') as fileWriter:
        fileWriter.write(pyCode)

    return (output_file,RegBankMes['name'])
                        
# print(CreatPy())